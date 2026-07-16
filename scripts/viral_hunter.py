#!/usr/bin/env python3
"""
爆款识别 + detail补抓命令生成 — 两阶段抓取方法（方案B），抖音/小红书双平台支持
"""
import sys, io
if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'buffer'):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import os, json, glob, argparse

# ── 平台差异化 ──
PLATFORM_DEFAULTS = {
    "dy": {
        "id_field": "aweme_id",
        "profile_tpl": "https://www.douyin.com/user/{uid}",
        "search_c": "search_comments_*.jsonl",
        "search_v": "search_contents_*.jsonl",
        "detail_c": "detail_comments_*.jsonl",
        "detail_v": "detail_contents_*.jsonl",
        "like_thresh": 10000,
        "comment_thresh": 1000,
        "detail_comments": 300,
    },
    "xhs": {
        "id_field": "note_id",
        "profile_tpl": "https://www.xiaohongshu.com/user/profile/{uid}",
        "search_c": "search_comments_*.jsonl",
        "search_v": "search_contents_*.jsonl",
        "detail_c": "detail_comments_*.jsonl",
        "detail_v": "detail_contents_*.jsonl",
        "like_thresh": 5000,
        "comment_thresh": 200,
        "detail_comments": 200,
    },
}


def detect_platform(path):
    """从路径自动识别平台"""
    if "xhs" in path.lower():
        return "xhs"
    return "dy"


def cfg(args):
    """获取当前平台配置"""
    plat = args.platform or "dy"
    return PLATFORM_DEFAULTS[plat], plat


def load(pattern):
    rows = []
    for p in glob.glob(pattern):
        with open(p, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
    return rows


def parse_count(v):
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if not s:
        return 0
    try:
        return int(s)
    except ValueError:
        if "万" in s:
            try:
                return int(float(s.replace("万", "")) * 10000)
            except ValueError:
                return 0
        return 0


def find_viral(works, id_field, like_thresh, comment_thresh):
    viral, seen = [], set()
    for w in works:
        aid = w.get(id_field, "")
        if aid in seen:
            continue
        seen.add(aid)
        likes = parse_count(w.get("liked_count", 0))
        cmts = parse_count(w.get("comment_count", 0))
        if likes >= like_thresh or cmts >= comment_thresh:
            viral.append((likes, cmts, w))
    viral.sort(key=lambda x: -x[0])
    return viral


def show_viral(viral, id_field, like_thresh, comment_thresh):
    label = "视频" if id_field == "aweme_id" else "笔记"
    print(f"爆款{label}: {len(viral)} 个 (点赞>{like_thresh} 或 评论>{comment_thresh})\n")
    for i, (l, c, w) in enumerate(viral, 1):
        title = (w.get("title") or w.get("desc") or "")[:55]
        print(f"{i:2}. [{l}赞 / {c}评] {title}")
        print(f"    id: {w[id_field]}")
    return [w[id_field] for _, _, w in viral]


def gen_command(platform, ids, max_comments):
    print(f"\n=== {platform.upper()} detail 补抓命令 ===\n")
    print(f"# 补抓 {len(ids)} 个爆款，每篇 {max_comments} 条评论")
    spec = "specified_id" if platform == "dy" else "specified_id"
    cmd = (
        f'uv run main.py --platform {platform} --lt qrcode --type detail '
        f'--{spec} "{"，".join(ids)}" '
        f'--max_comments_count_singlenotes {max_comments}'
    )
    print(cmd)
    return cmd


def merge_and_export(data_dir, out_path, platform):
    plat_cfg = PLATFORM_DEFAULTS.get(platform, PLATFORM_DEFAULTS["dy"])
    id_f = plat_cfg["id_field"]
    prof_t = plat_cfg["profile_tpl"]
    data_dir = data_dir.rstrip("/\\")

    sc = load(os.path.join(data_dir, plat_cfg["search_c"]))
    sv = load(os.path.join(data_dir, plat_cfg["search_v"]))
    dc = load(os.path.join(data_dir, plat_cfg["detail_c"]))
    dv = load(os.path.join(data_dir, plat_cfg["detail_v"]))

    # 去重
    v_seen, videos = set(), []
    for d in sv + dv:
        aid = d.get(id_f)
        if aid and aid not in v_seen:
            v_seen.add(aid)
            videos.append(d)

    c_dict = {}
    for d in sc:
        cid = d.get("comment_id")
        if cid:
            c_dict[cid] = d
    detail_from = 0
    for d in dc:
        cid = d.get("comment_id")
        if cid:
            c_dict[cid] = d
            detail_from += 1

    comments = list(c_dict.values())
    label = "视频" if platform == "dy" else "笔记"
    print(f"合并: {label} {len(videos)} / 评论 {len(comments)} (detail来源 {detail_from})")

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HF = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    HFT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    BF = Font(name="微软雅黑", size=10)
    B = Border(left=Side(style="thin"), right=Side(style="thin"),
               top=Side(style="thin"), bottom=Side(style="thin"))

    wb = Workbook()
    ws = wb.active; ws.title = "评论"
    headers = ["序号","评论ID","所属ID","评论内容","点赞","IP属地","用户ID","昵称","用户主页","来源"]
    ws.append(headers)
    for ci in range(1, len(headers)+1):
        c = ws.cell(row=1, column=ci); c.fill = HF; c.font = HFT; c.border = B
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    detail_ids = {d[id_f] for d in dv}
    for i, cm in enumerate(comments, 1):
        uid = cm.get("user_id") or cm.get("sec_uid") or ""
        src = "detail" if cm.get(id_f) in detail_ids else "search"
        vals = [i, cm.get("comment_id"), cm.get(id_f), cm.get("content"),
                cm.get("like_count"), cm.get("ip_location"), uid, cm.get("nickname"),
                prof_t.format(uid=uid) if uid else "", src]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=i+1, column=ci, value=v); c.font = BF; c.border = B
    ws.column_dimensions["D"].width = 55; ws.column_dimensions["I"].width = 42
    wb.save(out_path)
    print(f"导出完成: {out_path}")


def main():
    ap = argparse.ArgumentParser(description="爆款识别 + detail补抓 — 两阶段(抖音/小红书)")
    ap.add_argument("path", help="search_contents jsonl路径或含子文件的目录")
    ap.add_argument("--platform", choices=["dy","xhs"], help="平台(dy/xhs),不指定则自动识别")
    ap.add_argument("--command", action="store_true", help="输出可执行的 detail 补抓命令")
    ap.add_argument("--merge", action="store_true", help="合并 search+detail 并导出xlsx")
    ap.add_argument("--out", default="merged.xlsx", help="合并导出路径")
    ap.add_argument("--detail-comments", type=int, default=None, help="detail 评论上限")
    ap.add_argument("--like-thresh", type=int, default=None, help="爆款点赞阈值")
    ap.add_argument("--comment-thresh", type=int, default=None, help="爆款评论数阈值")
    args = ap.parse_args()

    if args.platform is None:
        args.platform = detect_platform(args.path)

    plat_cfg, platform = cfg(args)
    id_f = plat_cfg["id_field"]

    if args.merge:
        merge_and_export(args.path, args.out, platform)
        return

    works = load(args.path)
    label = "视频" if platform == "dy" else "笔记"
    print(f"{label}总数(去重): {len(works)}")

    viral = find_viral(
        works, id_f,
        args.like_thresh or plat_cfg["like_thresh"],
        args.comment_thresh or plat_cfg["comment_thresh"],
    )
    ids = show_viral(
        viral, id_f,
        args.like_thresh or plat_cfg["like_thresh"],
        args.comment_thresh or plat_cfg["comment_thresh"],
    )

    if args.command and ids:
        gen_command(platform, ids, args.detail_comments or plat_cfg["detail_comments"])


if __name__ == "__main__":
    main()
