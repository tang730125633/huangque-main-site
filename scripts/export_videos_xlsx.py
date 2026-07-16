#!/usr/bin/env python3
"""
视频级数据导出器 —— 把 MediaCrawler 的 search_contents jsonl 导出为 xlsx，
对齐「抖音医美数据.xlsx」的 18 列格式（视频维度全字段）。

增强（命令行接口向后兼容）：
- 表头配色 + 边框 + 冻结首行 + 长文本自动换行
- 视频链接/用户主页/封面/视频下载为「可点击超链接」
- 可选第 3 个参数传入 search_comments_*.jsonl，则额外导出「评论」sheet
  (含用户主页超链接)，关键词获客时一并拿到评论区线索原文

用法：
  python export_videos_xlsx.py <search_contents_*.jsonl> <输出.xlsx> [search_comments_*.jsonl]
"""
import sys
import glob
import json
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CST = timezone(timedelta(hours=8))
HEADERS = ["序号", "视频ID", "标题", "描述", "发布时间", "用户ID", "抖音号", "昵称",
           "点赞", "收藏", "评论数", "分享", "IP属地", "视频链接", "用户主页",
           "封面", "视频下载", "来源关键词"]
CMT_HEADERS = ["序号", "评论内容", "昵称", "抖音号", "属地", "点赞", "所属视频ID", "用户主页"]

HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
LINK_FONT = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
BODY_FONT = Font(name="微软雅黑", size=10)
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fmt_time(ts):
    try:
        return datetime.fromtimestamp(int(ts), CST).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


def load(pattern):
    rows = []
    for p in glob.glob(pattern):
        with open(p, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
    return rows


def style_header(ws, ncols):
    for ci in range(1, ncols + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.freeze_panes = "A2"


def put(ws, row, col, value, link=False, wrap=False):
    c = ws.cell(row=row, column=col)
    if link and value:
        c.value = value
        c.hyperlink = value
        c.font = LINK_FONT
    else:
        c.value = value
        c.font = BODY_FONT
    c.border = BORDER
    if wrap:
        c.alignment = Alignment(wrap_text=True)
    return c


def main():
    pattern, out = sys.argv[1], sys.argv[2]
    comments_pattern = sys.argv[3] if len(sys.argv) > 3 else ""
    rows = load(pattern)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "抖音视频"
    ws.append(HEADERS)
    style_header(ws, len(HEADERS))
    for i, d in enumerate(rows, 1):
        sec = d.get("sec_uid") or ""
        home = f"https://www.douyin.com/user/{sec}" if sec else ""
        vals = [i, d.get("aweme_id"), d.get("title"), d.get("desc"), fmt_time(d.get("create_time")),
                d.get("user_id"), d.get("user_unique_id"), d.get("nickname"),
                d.get("liked_count"), d.get("collected_count"), d.get("comment_count"),
                d.get("share_count"), d.get("ip_location"), d.get("aweme_url"),
                home, d.get("cover_url"), d.get("video_download_url"), d.get("source_keyword")]
        link_cols = {14, 15, 16, 17}
        wrap_cols = {3, 4}
        for ci, v in enumerate(vals, 1):
            put(ws, i + 1, ci, v, link=(ci in link_cols), wrap=(ci in wrap_cols))
    for col, w in {"C": 40, "D": 40, "N": 36, "O": 36, "P": 30, "Q": 30}.items():
        ws.column_dimensions[col].width = w

    # 可选：评论 sheet
    cmt_count = 0
    if comments_pattern:
        comments = load(comments_pattern)
        cmt_count = len(comments)
        ws2 = wb.create_sheet("评论")
        ws2.append(CMT_HEADERS)
        style_header(ws2, len(CMT_HEADERS))
        for i, cm in enumerate(comments, 1):
            sec = cm.get("sec_uid") or ""
            home = f"https://www.douyin.com/user/{sec}" if sec else ""
            vals = [i, cm.get("content"), cm.get("nickname"), cm.get("user_unique_id"),
                    cm.get("ip_location"), cm.get("like_count"), cm.get("aweme_id"), home]
            for ci, v in enumerate(vals, 1):
                put(ws2, i + 1, ci, v, link=(ci == 8), wrap=(ci == 2))
        ws2.column_dimensions["B"].width = 50
        ws2.column_dimensions["H"].width = 42

    wb.save(out)
    if comments_pattern:
        print(f"✅ 导出 {len(rows)} 条视频 + {cmt_count} 条评论 → {out}")
    else:
        print(f"✅ 导出 {len(rows)} 条视频 → {out}")


if __name__ == "__main__":
    main()
