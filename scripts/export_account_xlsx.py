#!/usr/bin/env python3
"""
账号 xlsx 导出器 —— 把 MediaCrawler creator 模式的输出导成「一个账号一个 xlsx」。
含 4 个 sheet：① 账号画像 ② 全部作品(18列) ③ 评论 ④ 评论统计(各视频评论数)。

增强（保持命令行接口不变，worker.py 调用方式不受影响）：
- 表头配色 + 边框 + 冻结首行 + 长文本自动换行
- 主页/视频/封面/下载/视频链接为「可点击超链接」(cell.hyperlink)
- 评论 sheet 增加「用户主页」列
- 按账号 sec_uid 过滤 + aweme_id/comment_id 去重(解决多账号 jsonl 追加混写)
- 新增「评论统计」sheet：各视频抓到多少评论

用法：python export_account_xlsx.py <jsonl_dir> <输出.xlsx>
"""
import sys
import os
import glob
import json
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CST = timezone(timedelta(hours=8))
WORK_HEADERS = ["序号", "视频ID", "标题", "描述", "发布时间", "用户ID", "抖音号", "昵称",
                "点赞", "收藏", "评论数", "分享", "IP属地", "视频链接", "用户主页",
                "封面", "视频下载", "来源关键词"]
CMT_HEADERS = ["序号", "评论内容", "昵称", "抖音号", "属地", "点赞", "所属视频ID", "用户主页"]
STAT_HEADERS = ["视频ID", "标题", "评论抓取数", "视频链接"]

# 样式
HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
LINK_FONT = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
BODY_FONT = Font(name="微软雅黑", size=10)
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load(pattern):
    rows = []
    for p in glob.glob(pattern):
        with open(p, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
    return rows


def ftime(ts):
    try:
        return datetime.fromtimestamp(int(ts), CST).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


def style_header(ws, ncols):
    """给首行表头上色 + 边框，并冻结首行。"""
    for ci in range(1, ncols + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws.freeze_panes = "A2"


def put(ws, row, col, value, link=False, wrap=False):
    """写单元格并套用样式；link=True 时写为可点击超链接。"""
    c = ws.cell(row=row, column=col)
    if link and value:
        c.value = value
        c.hyperlink = value
        c.font = LINK_FONT
    else:
        c.value = value
        c.font = BODY_FONT
    c.border = BORDER
    if wrap:
        c.alignment = Alignment(wrap_text=True)
    return c


def main():
    jdir, out = sys.argv[1], sys.argv[2]
    creators = load(os.path.join(jdir, "creator_creators_*.jsonl"))
    works_all = load(os.path.join(jdir, "creator_contents_*.jsonl"))
    comments_all = load(os.path.join(jdir, "creator_comments_*.jsonl"))

    # 目标账号 sec_uid：以抓到的作品里出现的为准(jsonl 可能追加混入了别的账号)
    sec = works_all[0].get("sec_uid") if works_all else ""

    # 按 sec_uid 过滤 + aweme_id 去重
    works, seen_aweme, target_aweme = [], set(), set()
    for w in works_all:
        if sec and w.get("sec_uid") != sec:
            continue
        aid = w.get("aweme_id")
        if aid in seen_aweme:
            continue
        seen_aweme.add(aid)
        target_aweme.add(aid)
        works.append(w)

    # 评论：只留目标账号视频下的 + comment_id 去重
    comments, seen_cmt = [], set()
    for c in comments_all:
        if target_aweme and c.get("aweme_id") not in target_aweme:
            continue
        cid = c.get("comment_id")
        if cid and cid in seen_cmt:
            continue
        seen_cmt.add(cid)
        comments.append(c)

    wb = openpyxl.Workbook()

    # ---- sheet1 画像 ----
    ws = wb.active
    ws.title = "账号画像"
    c = creators[0] if creators else {}
    ip = (c.get("ip_location") or "").replace("IP属地：", "")
    ws.append(["项目", "数据"])
    style_header(ws, 2)
    profile_url = f"https://www.douyin.com/user/{sec}" if sec else ""
    rows = [
        ("昵称", c.get("nickname")),
        ("抖音号", works[0].get("user_unique_id") if works else ""),
        ("粉丝", c.get("fans")), ("获赞", c.get("interaction")),
        ("作品数", c.get("videos_count")), ("关注", c.get("follows")),
        ("IP属地", ip), ("简介", c.get("desc") or "无"),
        ("主页", profile_url), ("头像", c.get("avatar") or ""),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        kc = ws.cell(row=i, column=1, value=k)
        kc.font = Font(name="微软雅黑", size=10, bold=True)
        kc.border = BORDER
        put(ws, i, 2, v, link=(k in ("主页", "头像")), wrap=(k == "简介"))
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60

    # ---- sheet2 作品 ----
    ws2 = wb.create_sheet("全部作品")
    ws2.append(WORK_HEADERS)
    style_header(ws2, len(WORK_HEADERS))
    for i, d in enumerate(works, 1):
        s = d.get("sec_uid") or ""
        home = f"https://www.douyin.com/user/{s}" if s else ""
        vals = [i, d.get("aweme_id"), d.get("title"), d.get("desc"), ftime(d.get("create_time")),
                d.get("user_id"), d.get("user_unique_id"), d.get("nickname"),
                d.get("liked_count"), d.get("collected_count"), d.get("comment_count"),
                d.get("share_count"), d.get("ip_location"), d.get("aweme_url"),
                home, d.get("cover_url"), d.get("video_download_url"), d.get("source_keyword")]
        link_cols = {14, 15, 16, 17}   # 视频链接/用户主页/封面/视频下载
        wrap_cols = {3, 4}             # 标题/描述
        for ci, v in enumerate(vals, 1):
            put(ws2, i + 1, ci, v, link=(ci in link_cols), wrap=(ci in wrap_cols))
    for col, w in {"C": 40, "D": 40, "N": 34, "O": 34, "P": 30, "Q": 30}.items():
        ws2.column_dimensions[col].width = w

    # ---- sheet3 评论 ----
    ws3 = wb.create_sheet("评论")
    ws3.append(CMT_HEADERS)
    style_header(ws3, len(CMT_HEADERS))
    for i, cm in enumerate(comments, 1):
        s = cm.get("sec_uid") or ""
        home = f"https://www.douyin.com/user/{s}" if s else ""
        vals = [i, cm.get("content"), cm.get("nickname"), cm.get("user_unique_id"),
                cm.get("ip_location"), cm.get("like_count"), cm.get("aweme_id"), home]
        for ci, v in enumerate(vals, 1):
            put(ws3, i + 1, ci, v, link=(ci == 8), wrap=(ci == 2))
    ws3.column_dimensions["B"].width = 50
    ws3.column_dimensions["H"].width = 42

    # ---- sheet4 评论统计 ----
    ws4 = wb.create_sheet("评论统计")
    ws4.append(STAT_HEADERS)
    style_header(ws4, len(STAT_HEADERS))
    title_map = {w.get("aweme_id"): (w.get("title") or w.get("desc") or "")[:30] for w in works}
    stat = {}
    for cm in comments:
        stat[cm.get("aweme_id")] = stat.get(cm.get("aweme_id"), 0) + 1
    for i, (aid, cnt) in enumerate(sorted(stat.items(), key=lambda x: -x[1]), 1):
        url = f"https://www.douyin.com/video/{aid}" if aid else ""
        for ci, v in enumerate([aid, title_map.get(aid, ""), cnt, url], 1):
            put(ws4, i + 1, ci, v, link=(ci == 4), wrap=(ci == 2))
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 40
    ws4.column_dimensions["D"].width = 48

    wb.save(out)
    print(f"✅ 账号「{c.get('nickname')}」→ {out}")
    print(f"   画像: 粉丝{c.get('fans')}/获赞{c.get('interaction')}/作品{c.get('videos_count')} | "
          f"作品表 {len(works)} 条 | 评论 {len(comments)} 条")


if __name__ == "__main__":
    main()
