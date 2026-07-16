#!/usr/bin/env python3
"""
小红书数据导出器 — 把 MediaCrawler 的 search_contents/creator_contents jsonl 导出为 xlsx。
含 3 个 sheet: 笔记 / 评论 / 摘要。

和抖音导出脚本的关键差异: 字段名不同 —
  video_url(非 video_download_url), image_list(非 cover_url),
  note_id(非 aweme_id), user_id(非 sec_uid)。
用户主页: https://www.xiaohongshu.com/user/profile/{user_id}

用法:
  python scripts/export_xhs_xlsx.py <search_contents_*.jsonl> <search_comments_*.jsonl> <输出.xlsx>
"""
import sys, os, glob, json
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CST = timezone(timedelta(hours=8))
NOTE_HEADERS = ["序号","笔记ID","类型","标题","描述","发布时间","用户ID","昵称",
                "点赞","收藏","评论数","分享","IP属地","标签","笔记链接","用户主页",
                "视频下载","封面图片"]
CMT_HEADERS = ["序号","评论内容","昵称","属地","点赞","所属笔记ID","用户ID","用户主页"]

HEADER_FILL = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
LINK_FONT = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
BODY_FONT = Font(name="微软雅黑", size=10)
THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load(pattern):
    rows = []
    for p in glob.glob(pattern):
        with open(p, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    rows.append(json.loads(line))
    return rows


def style_header(ws, ncols):
    for ci in range(1, ncols + 1):
        c = ws.cell(row=1, column=ci)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center"); c.border = BORDER
    ws.freeze_panes = "A2"


def put(ws, row, col, value, link=False, wrap=False):
    c = ws.cell(row=row, column=col)
    if link and value:
        c.value = value; c.hyperlink = value; c.font = LINK_FONT
    else:
        c.value = value; c.font = BODY_FONT
    c.border = BORDER
    if wrap:
        c.alignment = Alignment(wrap_text=True)


def main():
    if len(sys.argv) < 4:
        print("用法: python export_xhs_xlsx.py <search_contents_*.jsonl> <search_comments_*.jsonl> <输出.xlsx>")
        sys.exit(1)
    notes_pat, cmts_pat, out = sys.argv[1], sys.argv[2], sys.argv[3]
    notes = load(notes_pat)
    comments = load(cmts_pat)

    wb = openpyxl.Workbook()

    # ── Sheet1: 笔记 ──
    ws = wb.active; ws.title = "小红书笔记"
    ws.append(NOTE_HEADERS); style_header(ws, len(NOTE_HEADERS))
    for i, d in enumerate(notes, 1):
        uid = d.get("user_id", "")
        home = f"https://www.xiaohongshu.com/user/profile/{uid}" if uid else ""
        vals = [i, d.get("note_id"), d.get("type"), d.get("title"), d.get("desc"),
                d.get("time"), uid, d.get("nickname"),
                d.get("liked_count"), d.get("collected_count"), d.get("comment_count"),
                d.get("share_count"), d.get("ip_location"), d.get("tag_list"),
                d.get("note_url"), home, d.get("video_url", ""), d.get("image_list", "")]
        for ci, v in enumerate(vals, 1):
            put(ws, i+1, ci, v, link=(ci in {15,16,17}), wrap=(ci in {4,5}))
    for col, w in {"D":42,"E":42,"I":10,"J":10,"O":50,"P":42,"Q":45,"R":45}.items():
        ws.column_dimensions[col].width = w

    # ── Sheet2: 评论 ──
    ws2 = wb.create_sheet("小红书评论")
    ws2.append(CMT_HEADERS); style_header(ws2, len(CMT_HEADERS))
    for i, d in enumerate(comments, 1):
        uid = d.get("user_id", "")
        home = f"https://www.xiaohongshu.com/user/profile/{uid}" if uid else ""
        vals = [i, d.get("content"), d.get("nickname"), d.get("ip_location"),
                d.get("like_count"), d.get("note_id"), uid, home]
        for ci, v in enumerate(vals, 1):
            put(ws2, i+1, ci, v, link=(ci == 8), wrap=(ci == 2))
    ws2.column_dimensions["B"].width = 55; ws2.column_dimensions["H"].width = 42

    # ── Sheet3: 摘要 ──
    ws3 = wb.create_sheet("摘要")
    ws3.append(["项目","数据"]); style_header(ws3, 2)
    for k, v in [("笔记数", len(notes)), ("评论数", len(comments)), ("时间", datetime.now(CST).strftime("%Y-%m-%d"))]:
        ws3.append([k, v])
    ws3.column_dimensions["A"].width = 20

    wb.save(out)
    print(f"导出完成: {out}")
    print(f"  小红书笔记 {len(notes)} 条 / 评论 {len(comments)} 条")


if __name__ == "__main__":
    main()
